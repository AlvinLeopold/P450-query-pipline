#P450页面爬虫
# -*- coding: utf-8 -*-
# @Time    : 2021/3/22 10:30
# @Author  : AlvinLiu
# @Site    : Cytochrm P450
# @Email   : JakeLiu_ok@163.com (Don't hesitate to contact me if you have any questions)

import requests
from bs4 import BeautifulSoup
import time
import openpyxl
import re


start_time = time.time()

#Session functions are used to keep the login status
requests = requests.Session()

#Headers in necessary, you can find it in the browser's developer tools and replace ********* with your own
headers = {
    'User-Agent':'*********',
    'Referer':'http://passport.riceblast.snu.ac.kr/?t=P450'
}

def creat_excel(out_file, ori_path):
    wb_out = openpyxl.Workbook()
    sheet = wb_out.active
    sheet.title = 'Sheet1'
    sheet['A1'] = 'id'
    sheet['B1'] = 'query'
    sheet['C1'] = 'P450Name'
    sheet['D1'] = 'P450Classname'
    sheet['E1'] = 'PutativeName'
    sheet['F1'] = 'NelsonName'
    sheet['G1'] = 'Hit_species'
    wb_out.save(ori_path + out_file)
    return wb_out

# Visit the login page
def login(id, pw):
    data={
        'a': 'login',
        'r': 'http://p450.riceblast.snu.ac.kr/index.php',
        'b': 'snu.ac.kr',
        'id': id,
        'pw': pw
    }
    print('The program is running...logging in')
    login_response = requests.post('http://passport.riceblast.snu.ac.kr/login.php', data=data, headers=headers)
    print('The status code is 200 means it working'+str(login_response.status_code))
    return login_response.status_code

def submit(query):
    print('Looking up:', query)
    data = {
        'a': 'search',
        'sf': 'SEQUENCE_NAME',
        'sv': query
    }
    #Get the id of the query
    submit_response = requests.post('http://p450.riceblast.snu.ac.kr/index.php?a=search&sf=SEQUENCE_NAME&sv='+query, data=data, headers=headers)
    id_r = re.findall(r'id=(\d+)', submit_response.text)
    id_re = ''.join(id_r)
    return id_re

def get_result(id_re):
    print('Submitting the query...')
    data = {
        'a': 'dv_sequence',
        'id': id_re,
    }
    submit_response2 = requests.post('http://p450.riceblast.snu.ac.kr/class.php?a=dv_sequence&id=' + id,
                                    data=data, headers=headers)
    '''
    #将submit_response2保存为html文件
    with open('Path', 'w', encoding='utf-8') as f:
        f.write(submit_response2.text)
    '''

    #Get P450Name
    P450Name = re.findall(r'<td width="670">&nbsp;(.+?)\n\t\t\t\t\t\t</td>', submit_response2.text)
    P450Name_text = ''.join(P450Name)

    PutativeName = re.findall(r'<td width="70">&nbsp;(.+?)</td>', submit_response2.text)
    PutativeName_text = ''.join(PutativeName)

    #Get P450Classname, NelsonName, Hit_species, P450Name, PutativeName
    soup = BeautifulSoup(submit_response2.text, 'html.parser')
    first_link = soup.a
    wrapper = first_link.find_all_next(string=True)
    try:
        for j in range(127,141):
            if wrapper[j] == 'P450 Class Name':
                P450Classname_text = wrapper[j+6]
            else:
                pass

        for k in range(174,188):
            if wrapper[k] == 'Nelson\'s P450 Name':
                NelsonName_text = wrapper[k+17].strip('\n\t\t\t\t\t\t')
            else:
                pass

        if 'NelsonName_text' not in locals():
            NelsonName_text = None

        for s in range(202,222):
            if wrapper[s] == 'Species Name' and wrapper[s+6] != '\n':
                Hit_species_text = wrapper[s+6]
            else:
                pass

        if 'Hit_species_text' not in locals():
            Hit_species_text = None
        print('Result has been returned!')

    except:
        print('Negative result!')
        P450Classname_text = None
        NelsonName_text = None
        Hit_species_text = None
        P450Name_text = None
        PutativeName_text = None

    return P450Name_text, P450Classname_text, PutativeName_text, NelsonName_text, Hit_species_text


if __name__ == '__main__':

    print('程序开始运行！')

    id = input('请输入您的账号：/input your account:')
    pw = input('请输入您的密码：/input your password:')
    ori_path = input('请输入工作路径：./input your file path:')
    ori_file = input('请输入文件名：/input your file name:')
    out_file = ori_file.split('.')[0] + '_result.xlsx'
    file_path = ori_path + ori_file

    print('正在登录...')
    login(id, pw)
    print('登录成功！')
    #读入excel文件
    wb = openpyxl.load_workbook(file_path)
    sheet = wb['Sheet1']
    creat_excel(out_file, ori_path)
    row = sheet.max_row

    #遍历excel文件中的每一行
    for i in range(2, row+1):
        time.sleep(2)
        #通过tqdm库实现进度条进度=i/row
        # 获取查询序列
        query = sheet.cell(row=i, column=2).value
        # 获取id
        id = sheet.cell(row=i, column=1).value
        P450Name_text, P450Classname_text, PutativeName_text, NelsonName_text, Hit_species_text = get_result(submit(query))
        print('已检索到数据！正在写入文件......')
        wb_out = openpyxl.load_workbook(ori_path + out_file)
        sheet_out = wb_out['Sheet1']
        sheet_out.cell(row=i, column=1).value = id
        sheet_out.cell(row=i, column=2).value = query
        sheet_out.cell(row=i, column=3).value = P450Name_text
        sheet_out.cell(row=i, column=4).value = P450Classname_text
        sheet_out.cell(row=i, column=5).value = PutativeName_text
        sheet_out.cell(row=i, column=6).value = NelsonName_text
        sheet_out.cell(row=i, column=7).value = Hit_species_text
        print('正在保存文件......')
        wb_out.save(ori_path + out_file)
        print('保存成功！已写入：'+ str(i) +'条数据')
        progress = round(i/row,2)*100
        print('-----------当前进度'+str(progress)+'%-------------------')

end = time.time()
print('程序运行时间：' + str((end - start_time)/360) + 'h')